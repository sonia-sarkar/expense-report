import re
from datetime import datetime
from pathlib import Path
import cv2
import pytesseract
from openpyxl import Workbook, load_workbook
import csv
import os

# --------- Paths ---------
RECEIPT_DIR = Path("./receipts")
EXCEL_FILE = Path("./expenses.xlsx")
CSV_FILE = Path("./expenses.csv")

# Adjust this path if your tesseract is installed elsewhere
pytesseract.pytesseract.tesseract_cmd = "/opt/homebrew/bin/tesseract"

# --------- OCR and parsing functions ---------
def extract_text(image_path):
    """Reads image with OpenCV, converts to grayscale, and extracts text with Tesseract"""
    image = cv2.imread(str(image_path))
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    text = pytesseract.image_to_string(gray, config="--psm 6 --oem 3")
    return text

def extract_vendor(text):
    """Assumes vendor is the first non-empty line"""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return lines[0] if lines else "Unknown"

def extract_date(text):
    """Finds the first date in the text and returns it in YYYY-MM-DD format"""
    date_match = re.search(r'\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b', text)
    if not date_match:
        return ""
    date_str = date_match.group(1)
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"):
        try:
            return datetime.strptime(date_str, fmt).date().isoformat()
        except ValueError:
            continue
    return date_str  # fallback to raw match

def extract_amount(text):
    """Finds the first monetary amount in the text (e.g., $12.34)"""
    match = re.search(r'\$?\s*([0-9]+\.[0-9]{2})', text)
    return match.group(1) if match else ""

# --------- File writing functions ---------
def save_to_excel(date, vendor, amount, notes, filename=EXCEL_FILE):
    """Append a row to Excel, create file if it doesn't exist"""
    if filename.exists():
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Vendor", "Amount", "Notes"])
    ws.append([date, vendor, amount, notes])
    wb.save(filename)

def save_to_csv(date, vendor, amount, notes, filename=CSV_FILE):
    """Append a row to CSV, create file with header if needed"""
    if not filename.exists():
        with open(filename, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Date", "Vendor", "Amount", "Notes"])
    with open(filename, "a", newline="") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow([date, vendor, amount, notes])

# --------- Main processing ---------
def process_receipts():
    for receipt_path in RECEIPT_DIR.glob("*.jpg"):
        print(f"Processing {receipt_path.name}...")
        text = extract_text(receipt_path)
        vendor = extract_vendor(text)
        date = extract_date(text)
        amount = extract_amount(text)
        notes = "Receipt scanned automatically"

        save_to_excel(date, vendor, amount, notes)
        save_to_csv(date, vendor, amount, notes)
        print(f"Added: Date={date}, Vendor={vendor}, Amount={amount}\n")

if __name__ == "__main__":
    process_receipts()
